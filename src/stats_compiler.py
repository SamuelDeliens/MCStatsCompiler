import json
import os
import pandas as pd
import numpy as np
import math
import datetime
import warnings
import openpyxl


def load_vanilla_data(csvtoggle, csvpath, inputmode, localpath):
    """Load and process vanilla Minecraft stats data"""
    df = pd.DataFrame()

    # Determine the source paths based on input mode
    if inputmode == "manual":
        names_file = open('data/usercache/usercache.json', 'r')
        stats_path = 'data/stats'
    elif inputmode == "local":
        names_file = open(localpath + '/usercache.json', 'r')
        stats_path = localpath + '/world/stats'
    else:
        raise ValueError(f"Unsupported input mode: {inputmode}")

    # Load the player names
    names = pd.DataFrame(json.load(names_file))

    # Process each stats file
    for filename in os.listdir(stats_path):
        if filename == ".gitignore":
            continue
        print("Now processing", filename)

        file = open(stats_path + '/' + filename)
        data = json.load(file)

        # Import the JSON to a Pandas DF
        temp_df = pd.json_normalize(data, meta_prefix=True)
        temp_name = names.loc[names['uuid'] == filename[:-5]]['name']

        if temp_name.empty:
            print("No username found for UUID", filename[:-5],
                  " in usercache.json, using UUID for this player instead.")
            temp_name = filename[:-5]
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name}, axis=1)
        else:
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)

        # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
        temp_df.index = temp_df.index.str.split('.', expand=True)

        # If a stat name has a dot in it, remove the part after the dot
        if len(temp_df.index.levshape) > 3:
            temp_df.index = temp_df.index.droplevel(3)
            temp_df = temp_df.groupby(level=[0, 1, 2]).sum()

        if df.empty:
            df = temp_df
        else:
            df = df.join(temp_df, how="outer")

    # Replace missing values by 0
    df = df.fillna(0)

    if csvtoggle == "true":
        df.to_csv(csvpath)

    return df


def load_cobblemon_data(csvtoggle, csvpath, inputmode, localpath):
    """Load and process Cobblemon mod data"""
    df = pd.DataFrame()
    root_dirnames = []

    # Determine the source paths based on input mode
    if inputmode == "manual":
        names_file = open('data/usercache/usercache.json', 'r')
        path = 'data/cobblemonplayerdata'
    elif inputmode == "local":
        names_file = open(localpath + '/usercache.json', 'r')
        path = localpath + '/world/cobblemonplayerdata'
    else:
        raise ValueError(f"Unsupported input mode: {inputmode}")

    # Load the player names
    names = pd.DataFrame(json.load(names_file))

    # Process each Cobblemon data file
    i = -1
    for dirpath, dirnames, filenames in os.walk(path):
        if len(dirnames) > 0:
            root_dirnames = dirnames
        for filename in filenames:
            if filename == ".gitignore" or filename == ".DS_Store":
                continue
            print("Now processing", filename)

            file = open(path + '/' + root_dirnames[i] + '/' + filename)
            data = json.load(file)['extraData']['cobbledex_discovery']['registers']

            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            temp_df = temp_df.transpose().iloc[:]

            if temp_name.empty:
                print("No username found for UUID", filename[:-5],
                      " in usercache.json, using UUID for this player instead.")
                temp_name = filename[:-5]
                temp_df = temp_df.rename({0: temp_name}, axis=1)
            else:
                temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)

            # Split the index
            if not temp_df.empty:
                temp_df.index = temp_df.index.str.split('.', expand=True)
                if df.empty:
                    df = temp_df
                else:
                    df = df.join(temp_df, how="outer")
            else:
                df[temp_name] = np.nan
        i += 1

    # Replace missing values by 0
    df = df.fillna(0)

    if csvtoggle == "true":
        df.to_csv(csvpath)

    return df


def get_vanilla_leaderboard(df, cat, subcat):
    """Display a leaderboard for a specific vanilla Minecraft statistic"""
    row = df.loc['stats'].loc[cat].loc[subcat].sort_values(ascending=False)
    print(f"Leaderboard of {cat} {subcat}:")
    print(row)
    return row


def get_vanilla_best_and_worst(df, username, cleaning, cleaningvalue):
    """Display best and worst statistics for a specific player"""
    if username == "null" or not username:
        print("Error for Best-and-Worst feature: no username specified in the config")
        return

    if username not in df.columns:
        print(f"Error for Best-and-Worst feature: User '{username}' does not exist in the provided data")
        print("Available users:", ", ".join(df.columns))
        return

    nb_players = df.shape[1]

    # Clean data if required
    if cleaning == "true":
        before_value = df.shape[0]
        df['zero_count'] = df.apply(lambda row: (row == 0).sum(), axis=1)
        df.drop(df[df['zero_count'] > (nb_players - int(cleaningvalue))].index, inplace=True)
        df = df.drop('zero_count', axis=1)
        print(before_value - df.shape[0], "rows dropped out of", before_value, "because of cleaning.")

    # Compute ranks
    ranks = df.rank(axis=1, method='min', ascending=False)
    ranks['non_zero_values'] = df.apply(lambda row: nb_players - (row == 0).sum(), axis=1)
    ranks['value'] = df[username]

    # Format output
    output = ranks[[username, 'value', 'non_zero_values']].sort_values(
        username, ascending=False
    ).rename(columns={username: "rank_" + username, "value": "value_" + username})

    print(output)
    return output


def update_leaderboard_in_excel(df, config, type):
    """Update a leaderboard in the Excel output file"""
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)

    tab_map = {
        "standard": "leaderboard2",
        "shiny": "leaderboard3",
        "legendary": "leaderboard4"
    }

    ws = wb[tab_map[type]]
    i = 0
    ExcelRows = int(config['COBBLEMONLEADERBOARDS']['ExcelRows'])
    ExcelCols = int(config['COBBLEMONLEADERBOARDS']['ExcelColumns'])

    # Update the cells with player data
    for index, row in df[0:ExcelRows * ExcelCols].iterrows():
        ws.cell(row=(i % ExcelRows) + 3, column=2 + math.floor(i / ExcelRows) * 3, value=str(i + 1) + ".")
        ws.cell(row=(i % ExcelRows) + 3, column=3 + math.floor(i / ExcelRows) * 3, value=index)
        ws.cell(row=(i % ExcelRows) + 3, column=4 + math.floor(i / ExcelRows) * 3, value=row[0])
        i += 1

    # Update timestamp and subtitle
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows + 3, column=2, value=now.strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated']))
    ws.cell(row=ExcelRows + 4, column=2, value=config['COBBLEMONLEADERBOARDS']['Subtitle'])

    wb.save(file_path)


def update_leaderboard_in_database(df, config, type, conn):
    """Update a leaderboard in the SQLite database"""
    if conn is None:
        return

    cursor = conn.cursor()

    # Table selection by leaderboard type
    table_map = {
        "standard": "standard_leaderboard",
        "shiny": "shiny_leaderboard",
        "legendary": "legendary_leaderboard"
    }

    table_name = table_map[type]

    # Clear old data
    cursor.execute(f"DELETE FROM {table_name}")

    # New data insertion
    now = datetime.datetime.now().strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated'])
    for index, row in df.iterrows():
        cursor.execute(f'''
            INSERT INTO {table_name} (rank, player_name, score, last_updated)
            VALUES (?, ?, ?, ?)
        ''', (int(row['index']), index, int(row[0]), now))

    conn.commit()


def most_pokemons_leaderboard(df, config, type, conn):
    """Process and update a Pokemon leaderboard in both Excel and SQLite"""
    # Update SQLite if enabled
    if config['COBBLEMONLEADERBOARDS']['SQLiteOutput'] == "true" and conn is not None:
        update_leaderboard_in_database(df, config, type, conn)

    # Update Excel if enabled
    if config['COBBLEMONLEADERBOARDS']['XLSXOutput'] == "true":
        update_leaderboard_in_excel(df, config, type)


def process_cobblemon_leaderboards(cobblemon_df, config, conn):
    """Process all Cobblemon leaderboards based on configuration"""
    # Prepare the counting DF
    count_df = cobblemon_df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
    pokemons_db = pd.read_csv('Pokemon.csv')
    legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

    # Parse ignore names list
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]

    # Total leaderboard feature
    if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
        player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
        player_sum['index'] = range(len(player_sum), 0, -1)
        player_sum = player_sum.iloc[::-1]
        player_sum.drop(ignore_names, inplace=True, errors='ignore')
        most_pokemons_leaderboard(player_sum, config, "standard", conn)

    # Shiny leaderboard feature
    if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
        player_sum = pd.DataFrame(((cobblemon_df == "True") | (cobblemon_df == True)).sum().sort_values())
        player_sum['index'] = range(len(player_sum), 0, -1)
        player_sum = player_sum.iloc[::-1]
        player_sum.drop(ignore_names, inplace=True, errors='ignore')
        most_pokemons_leaderboard(player_sum, config, "shiny", conn)

    # Legendary leaderboard feature
    if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
        legs = legendary_list['Cobblemon'].tolist()
        leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]

        with warnings.catch_warnings():
            warnings.simplefilter(action='ignore', category=FutureWarning)
            leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)

        player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
        player_sum['index'] = range(len(player_sum), 0, -1)
        player_sum = player_sum.iloc[::-1]
        player_sum.drop(ignore_names, inplace=True, errors='ignore')
        most_pokemons_leaderboard(player_sum, config, "legendary", conn)